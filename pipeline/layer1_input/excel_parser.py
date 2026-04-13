"""
Excel/CSV 解析器
从XLSX/CSV文件提取表格数据，转换为RawContent
"""

import csv
import os
from pathlib import Path
from models import RawContent, TableData


class ExcelParser:
    """Excel/CSV文件解析器"""

    MAX_ROWS = 1000  # 单sheet最多读取行数

    def parse(self, file_path: str) -> RawContent:
        """
        解析Excel/CSV文件为RawContent

        Args:
            file_path: 文件路径（.xlsx/.csv）

        Returns:
            RawContent with source_type="excel"
        """
        ext = Path(file_path).suffix.lower()

        if ext == ".xlsx":
            tables = self._parse_xlsx(file_path)
        elif ext == ".csv":
            tables = self._parse_csv(file_path)
        elif ext == ".xls":
            raise ValueError(
                "不支持旧版XLS格式，请用Excel或WPS转换为XLSX格式"
            )
        else:
            raise ValueError(f"不支持的文件格式: {ext}")

        raw_text = self._tables_to_text(tables)

        return RawContent(
            source_type="excel",
            raw_text=raw_text,
            tables=tables,
            metadata={
                "file_name": Path(file_path).name,
                "file_size": os.path.getsize(file_path),
                "table_count": len(tables),
                "format": ext,
            },
            detected_language=self._detect_language_from_tables(tables),
        )

    def _parse_xlsx(self, file_path: str) -> list:
        """解析XLSX文件"""
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise ImportError(
                "缺少openpyxl依赖，请运行: pip install openpyxl"
            )

        tables = []
        wb = load_workbook(file_path, read_only=True, data_only=True)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            rows_data = []
            for row in ws.iter_rows(values_only=True):
                # 将None转为空字符串，其他值保留
                cleaned = [
                    str(cell) if cell is not None else ""
                    for cell in row
                ]
                # 跳过全空行
                if any(c.strip() for c in cleaned):
                    rows_data.append(cleaned)

                if len(rows_data) >= self.MAX_ROWS:
                    break

            if not rows_data:
                continue

            headers = rows_data[0]
            data_rows = rows_data[1:]

            tables.append(TableData(
                headers=headers,
                rows=data_rows,
                source_sheet=sheet_name,
                source_range=f"A1:{chr(65 + len(headers) - 1)}{len(rows_data)}",
            ))

        wb.close()

        return tables

    def _parse_csv(self, file_path: str) -> list:
        """解析CSV文件"""
        tables = []

        # 尝试检测编码
        encoding = self._detect_csv_encoding(file_path)

        with open(file_path, 'r', encoding=encoding, errors='replace') as f:
            # 尝试自动检测方言
            sample = f.read(8192)
            f.seek(0)

            try:
                dialect = csv.Sniffer().sniff(sample)
            except csv.Error:
                dialect = csv.excel

            reader = csv.reader(f, dialect)
            rows_data = []
            for row in reader:
                cleaned = [cell.strip() for cell in row]
                if any(c for c in cleaned):
                    rows_data.append(cleaned)
                if len(rows_data) >= self.MAX_ROWS:
                    break

        if rows_data:
            headers = rows_data[0]
            data_rows = rows_data[1:]
            tables.append(TableData(
                headers=headers,
                rows=data_rows,
                source_sheet="CSV",
                source_range=f"A1:{chr(65 + len(headers) - 1)}{len(rows_data)}",
            ))

        return tables

    def _tables_to_text(self, tables: list) -> str:
        """
        将表格数据序列化为文本，供Layer 2消费

        格式：每张表用markdown table格式呈现
        """
        if not tables:
            return ""

        parts = []
        for table in tables:
            # 表头
            header_line = "| " + " | ".join(str(h) for h in table.headers) + " |"
            separator = "| " + " | ".join("---" for _ in table.headers) + " |"

            # 数据行
            data_lines = []
            for row in table.rows:
                cells = [str(v) if v is not None else "" for v in row]
                # 补齐列数
                while len(cells) < len(table.headers):
                    cells.append("")
                data_lines.append("| " + " | ".join(cells[:len(table.headers)]) + " |")

            sheet_label = f"[工作表: {table.source_sheet}]" if table.source_sheet else ""
            parts.append(
                f"{sheet_label}\n"
                f"{header_line}\n"
                f"{separator}\n"
                + "\n".join(data_lines)
            )

        return "\n\n".join(parts)

    def _detect_csv_encoding(self, file_path: str) -> str:
        """检测CSV文件编码"""
        # 简单检测：先试UTF-8，失败则试GBK
        for encoding in ['utf-8-sig', 'utf-8', 'gbk', 'gb18030']:
            try:
                with open(file_path, 'r', encoding=encoding) as f:
                    f.read(1024)
                return encoding
            except (UnicodeDecodeError, UnicodeError):
                continue
        return 'utf-8'  # 兜底

    def _detect_language_from_tables(self, tables: list) -> str:
        """从表格内容推断语言"""
        import re
        cjk_re = re.compile(r'[\u4e00-\u9fff]')
        all_text = ""
        for t in tables:
            all_text += " ".join(str(h) for h in t.headers)
            for row in t.rows[:5]:
                all_text += " ".join(str(v) for v in row if v)

        if not all_text:
            return "zh"

        cjk_count = len(cjk_re.findall(all_text))
        ratio = cjk_count / max(len(all_text), 1)
        return "zh" if ratio > 0.1 else "en"
